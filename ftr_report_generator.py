"""
Create Excel Template with Table formatting for Slicers
========================================================
Run this once to create the template, then add slicers manually in Excel.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path

TEMPLATE_PATH = Path('ftr_tracking/FTR_Report_Template.xlsx')
TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)

# Create workbook
wb = Workbook()

# =============================================================================
# Position_Summary Sheet
# =============================================================================
ws1 = wb.active
ws1.title = 'Position_Summary'

# Define columns
pos_columns = [
    'FTR_ID', 'Route', 'HedgeType', 'MW', 'Price_Paid', 'Owner',
    'Total_Settlement', 'Total_Cost', 'MTD_PnL', 'Days', 'Latest_Day_PnL', 'PnL_Per_MW'
]

# Write headers
for col_idx, col_name in enumerate(pos_columns, 1):
    ws1.cell(row=1, column=col_idx, value=col_name)

# Add sample data row (will be replaced by actual data)
sample_data = ['SAMPLE', 'OTA → HAY', 'OBL', 10, 5.5, 'Sample Owner', 1000, 500, 500, 30, 50, 50]
for col_idx, value in enumerate(sample_data, 1):
    ws1.cell(row=2, column=col_idx, value=value)

# Create Table (required for slicers)
table1 = Table(displayName="PositionSummary", ref=f"A1:L2")
table1.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
ws1.add_table(table1)

# Freeze header row
ws1.freeze_panes = 'A2'

# Set column widths
column_widths = {'A': 12, 'B': 15, 'C': 10, 'D': 8, 'E': 12, 'F': 25, 
                 'G': 15, 'H': 12, 'I': 12, 'J': 8, 'K': 15, 'L': 12}
for col, width in column_widths.items():
    ws1.column_dimensions[col].width = width

# =============================================================================
# Activity Sheet
# =============================================================================
ws2 = wb.create_sheet('Activity')

activity_columns = [
    'SnapshotDate', 'TransactionType', 'FTR_ID', 'Source', 'Sink',
    'HedgeType', 'MW_Previous', 'MW_Current', 'MW_Sold',
    'SaleProceeds', 'Profit', 'PricePerMW', 'CurrentOwner'
]

for col_idx, col_name in enumerate(activity_columns, 1):
    ws2.cell(row=1, column=col_idx, value=col_name)

# Sample data
sample_activity = ['2026-01-31', 'SELL', 'SAMPLE', 'OTA', 'HAY', 'OBL', 20, 10, 10, 5000, 2500, 10, 'Sample Owner']
for col_idx, value in enumerate(sample_activity, 1):
    ws2.cell(row=2, column=col_idx, value=value)

# Create Table
table2 = Table(displayName="ActivityLog", ref=f"A1:M2")
table2.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
ws2.add_table(table2)
ws2.freeze_panes = 'A2'

# =============================================================================
# Owner_Summary Sheet
# =============================================================================
ws3 = wb.create_sheet('Owner_Summary')

owner_columns = ['Owner', 'Total_Settlement', 'Total_Cost', 'MTD_PnL', 'Num_FTRs']

for col_idx, col_name in enumerate(owner_columns, 1):
    ws3.cell(row=1, column=col_idx, value=col_name)

sample_owner = ['Sample Owner', 10000, 5000, 5000, 10]
for col_idx, value in enumerate(sample_owner, 1):
    ws3.cell(row=2, column=col_idx, value=value)

table3 = Table(displayName="OwnerSummary", ref=f"A1:E2")
table3.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
ws3.add_table(table3)

# Save template
wb.save(TEMPLATE_PATH)
print(f"✓ Template created: {TEMPLATE_PATH}")
print("""
NEXT STEPS:
===========
1. Open the template in Excel: ftr_tracking/FTR_Report_Template.xlsx

2. Add Slicers:
   - Go to Position_Summary sheet
   - Click anywhere in the table
   - Insert → Slicer
   - Select: Owner, HedgeType, Route
   - Position slicers to the right of the data

3. Optional: Add slicers to Activity sheet for TransactionType, CurrentOwner

4. Save and close the template

The daily report script will now use this template and preserve your slicers!
""")
