"""
FTR Daily Report Generator
==========================
Generates daily Excel reports from the latest snapshot data.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from datetime import datetime
import logging

# =============================================================================
# CONFIGURATION
# =============================================================================

BASE_DIR = Path(__file__).parent / "ftr_tracking"
SNAPSHOT_DIR = BASE_DIR / "snapshots"
REPORTS_DIR = BASE_DIR / "reports"
LEDGER_FILE = BASE_DIR / "ftr_master_ledger.csv"

REPORTS_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# =============================================================================
# DATA LOADING
# =============================================================================

def get_latest_snapshot():
    """Find and load the most recent snapshot file."""
    snapshots = sorted(SNAPSHOT_DIR.glob("ftr_snapshot_*.csv"), reverse=True)
    if not snapshots:
        raise FileNotFoundError("No snapshot files found")
    
    latest = snapshots[0]
    date_str = latest.stem.replace("ftr_snapshot_", "")
    logger.info(f"Loading latest snapshot: {latest.name}")
    return pd.read_csv(latest), date_str


def get_previous_snapshot(current_date_str):
    """Load the previous day's snapshot for comparison."""
    current = datetime.strptime(current_date_str, '%Y%m%d')
    
    for days_back in range(1, 8):
        check_date = current - pd.Timedelta(days=days_back)
        check_str = check_date.strftime('%Y%m%d')
        filepath = SNAPSHOT_DIR / f"ftr_snapshot_{check_str}.csv"
        if filepath.exists():
            logger.info(f"Loading previous snapshot: {filepath.name}")
            return pd.read_csv(filepath), check_str
    
    return None, None


def load_ledger():
    """Load the master ledger."""
    if LEDGER_FILE.exists():
        return pd.read_csv(LEDGER_FILE)
    return pd.DataFrame()


# =============================================================================
# DATA PROCESSING
# =============================================================================
# DATA PROCESSING
# =============================================================================

def load_spot_prices_mtd(report_date_str):
    """Load spot prices for the month-to-date from cache."""
    try:
        # Get the month from report date (YYYYMMDD format)
        year_month = report_date_str[:6]  # YYYYMM
        spot_file = BASE_DIR / "spot_cache" / f"spot_{year_month}.csv"
        
        if not spot_file.exists():
            logger.warning(f"Spot price file not found: {spot_file}")
            return None, None
        
        spot_df = pd.read_csv(spot_file)
        spot_df['Trading date'] = pd.to_datetime(spot_df['Trading date'], dayfirst=True)
        
        # Get the report date (this is the date we're reporting FOR)
        # We want spot prices from the report date itself, not the day before
        report_date = datetime.strptime(report_date_str, '%Y%m%d')
        
        # Get first day of the month
        month_start = datetime(report_date.year, report_date.month, 1)
        
        # Filter to report date for daily P&L
        report_day_spot = spot_df[spot_df['Trading date'] == report_date].copy()
        
        # Filter to MTD (from month start to report date) for MTD P&L
        mtd_spot = spot_df[(spot_df['Trading date'] >= month_start) & 
                           (spot_df['Trading date'] <= report_date)].copy()
        
        # Calculate daily average for report date
        daily_avg = None
        if not report_day_spot.empty:
            daily_avg = report_day_spot.groupby('Point of connection')['$/MWh'].mean().reset_index()
            daily_avg['Node'] = daily_avg['Point of connection'].str.replace('2201', '')
        
        # Calculate MTD data for each node
        mtd_data = None
        if not mtd_spot.empty:
            # For obligations: simple average of price differences
            mtd_avg = mtd_spot.groupby('Point of connection')['$/MWh'].mean().reset_index()
            mtd_avg['Node'] = mtd_avg['Point of connection'].str.replace('2201', '')
            
            # For options: need to calculate max(0, diff) at half-hourly level then average
            # Group by date and node for daily calcs
            mtd_data = {
                'avg': mtd_avg,
                'spot_df': mtd_spot,
                'days_count': (report_date - month_start).days + 1
            }
        
        if daily_avg is not None:
            logger.info(f"Loaded spot prices - Report date: {report_date.date()}, MTD days: {mtd_data['days_count'] if mtd_data else 0}")
        
        return daily_avg, mtd_data
        
    except Exception as e:
        logger.error(f"Error loading spot prices: {e}")
        return None, None


def calculate_position_summary(df, report_date_str):
    """
    Calculate summary metrics for each FTR position.
    
    Uses the same logic as AllData notebook:
    - Options: max(0, sink - source) at each half-hourly trading period, then average
    - Obligations: (sink - source) simple average across trading periods
    - Profit: (settlement_price - price_paid) * MW * num_trading_periods * 0.5
    """
    # Load spot data
    daily_avg, mtd_data = load_spot_prices_mtd(report_date_str)
    
    # Get raw spot data for half-hourly calculations
    report_date = datetime.strptime(report_date_str, '%Y%m%d')
    month_start = datetime(report_date.year, report_date.month, 1)
    
    spot_df = None
    if mtd_data is not None:
        spot_df = mtd_data['spot_df']
    
    summary = []
    
    for _, row in df.iterrows():
        ftr_id = row['FTR_ID']
        source = row.get('Source', '')
        sink = row.get('Sink', '')
        route = f"{source} → {sink}"
        hedge_type = row.get('HedgeType', '')
        mw = float(row.get('MW', 0) or 0)
        price = float(row.get('Price', 0) or 0)
        owner = row.get('CurrentOwner', '')
        acq_cost = float(row.get('AcquisitionCost', 0) or 0)
        orig_cost = float(row.get('OriginalAcquisitionCost', 0) or 0)
        
        source_node = f"{source}2201"
        sink_node = f"{sink}2201"
        
        # Calculate days in contract and settlement period
        try:
            start = pd.to_datetime(row.get('StartDate'), dayfirst=True)
            end = pd.to_datetime(row.get('EndDate'), dayfirst=True)
            days = (end - start).days + 1
            settlement_period = start.strftime('%y%m')
        except:
            days = 0
            settlement_period = ''
        
        daily_pnl = 0
        mtd_pnl = 0
        
        if spot_df is not None:
            # --- Daily P&L (report date only) ---
            day_spot = spot_df[spot_df['Trading date'] == report_date]
            source_tp = day_spot[day_spot['Point of connection'] == source_node][['Trading period', '$/MWh']].copy()
            sink_tp = day_spot[day_spot['Point of connection'] == sink_node][['Trading period', '$/MWh']].copy()
            
            if not source_tp.empty and not sink_tp.empty:
                source_tp.columns = ['Trading period', 'Source_Price']
                sink_tp.columns = ['Trading period', 'Sink_Price']
                tp_merged = source_tp.merge(sink_tp, on='Trading period')
                tp_merged['Price_Diff'] = tp_merged['Sink_Price'] - tp_merged['Source_Price']
                
                if hedge_type == 'OPT':
                    tp_merged['Settlement'] = tp_merged['Price_Diff'].clip(lower=0)
                else:
                    tp_merged['Settlement'] = tp_merged['Price_Diff']
                
                daily_settlement = tp_merged['Settlement'].mean()
                daily_pnl = (daily_settlement - price) * mw * len(tp_merged) * 0.5
            
            # --- MTD P&L (all days from month start to report date) ---
            trading_dates = sorted(spot_df['Trading date'].unique())
            for trade_date in trading_dates:
                td_spot = spot_df[spot_df['Trading date'] == trade_date]
                src_tp = td_spot[td_spot['Point of connection'] == source_node][['Trading period', '$/MWh']].copy()
                snk_tp = td_spot[td_spot['Point of connection'] == sink_node][['Trading period', '$/MWh']].copy()
                
                if src_tp.empty or snk_tp.empty:
                    continue
                
                src_tp.columns = ['Trading period', 'Source_Price']
                snk_tp.columns = ['Trading period', 'Sink_Price']
                tp_m = src_tp.merge(snk_tp, on='Trading period')
                tp_m['Price_Diff'] = tp_m['Sink_Price'] - tp_m['Source_Price']
                
                if hedge_type == 'OPT':
                    tp_m['Settlement'] = tp_m['Price_Diff'].clip(lower=0)
                else:
                    tp_m['Settlement'] = tp_m['Price_Diff']
                
                day_settlement = tp_m['Settlement'].mean()
                mtd_pnl += (day_settlement - price) * mw * len(tp_m) * 0.5
        
        # Total settlement is the change in acquisition cost (what's actually been settled)
        total_settlement = orig_cost - acq_cost
        
        summary.append({
            'FTR_ID': ftr_id,
            'Settlement_Period': settlement_period,
            'Route': route,
            'HedgeType': hedge_type,
            'MW': mw,
            'Price_Paid': price,
            'Owner': owner,
            'Total_Settlement': total_settlement,
            'Total_Cost': orig_cost,
            'MTD_PnL': mtd_pnl,
            'Days': days,
            'Latest_Day_PnL': daily_pnl,
            'PnL_Per_MW': mtd_pnl / mw if mw > 0 else 0
        })
    
    return pd.DataFrame(summary)


def calculate_owner_summary(position_df):
    """Aggregate positions by owner."""
    if position_df.empty:
        return pd.DataFrame()
    
    owner_summary = position_df.groupby('Owner').agg({
        'Total_Settlement': 'sum',
        'Total_Cost': 'sum',
        'MTD_PnL': 'sum',
        'FTR_ID': 'count'
    }).reset_index()
    
    owner_summary.columns = ['Owner', 'Total_Settlement', 'Total_Cost', 'MTD_PnL', 'Num_FTRs']
    return owner_summary.sort_values('MTD_PnL', ascending=False)


def get_recent_activity(ledger_df, days=7):
    """Get recent trading activity from ledger."""
    if ledger_df.empty:
        return pd.DataFrame()
    
    ledger_df['SnapshotDate'] = pd.to_datetime(ledger_df['SnapshotDate'])
    cutoff = datetime.now() - pd.Timedelta(days=days)
    
    recent = ledger_df[ledger_df['SnapshotDate'] >= cutoff].copy()
    recent = recent[recent['TransactionType'] != 'INITIAL']
    
    return recent.sort_values('SnapshotDate', ascending=False)


# =============================================================================
# EXCEL REPORT GENERATION
# =============================================================================

def create_excel_report(position_df, activity_df, owner_df, report_date):
    """Create the Excel report with multiple sheets."""
    wb = Workbook()
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ==========================================================================
    # Position Summary Sheet
    # ==========================================================================
    ws1 = wb.active
    ws1.title = 'Position_Summary'
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(position_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
    
    # Create table for slicers
    if len(position_df) > 0:
        table_ref = f"A1:L{len(position_df) + 1}"
        table1 = Table(displayName="PositionSummary", ref=table_ref)
        table1.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws1.add_table(table1)
    
    ws1.freeze_panes = 'A2'
    
    # Column widths
    col_widths = [12, 15, 10, 8, 12, 20, 15, 12, 12, 8, 15, 12]
    for i, width in enumerate(col_widths, 1):
        ws1.column_dimensions[chr(64 + i)].width = width
    
    # ==========================================================================
    # Activity Sheet
    # ==========================================================================
    ws2 = wb.create_sheet('Activity')
    
    activity_cols = ['SnapshotDate', 'TransactionType', 'FTR_ID', 'Source', 'Sink',
                     'MW_Previous', 'MW_Current', 'MW_Sold', 'Notes']
    
    if not activity_df.empty:
        activity_export = activity_df[[c for c in activity_cols if c in activity_df.columns]].copy()
    else:
        activity_export = pd.DataFrame(columns=activity_cols)
    
    for r_idx, row in enumerate(dataframe_to_rows(activity_export, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    ws2.freeze_panes = 'A2'
    
    # ==========================================================================
    # Owner Summary Sheet
    # ==========================================================================
    ws3 = wb.create_sheet('Owner_Summary')
    
    for r_idx, row in enumerate(dataframe_to_rows(owner_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    ws3.freeze_panes = 'A2'
    
    # Save report
    report_filename = f"FTR_Daily_Report_{report_date}.xlsx"
    report_path = REPORTS_DIR / report_filename
    wb.save(report_path)
    logger.info(f"Report saved: {report_path}")
    
    return report_path


# =============================================================================
# EMAIL SUMMARY
# =============================================================================

def generate_email_summary(position_df, owner_df, report_date, report_filename):
    """Generate a text summary for email."""
    
    # Format date nicely
    try:
        date_obj = datetime.strptime(report_date, '%Y%m%d')
        date_formatted = date_obj.strftime('%B %d, %Y')
    except:
        date_formatted = report_date
    
    total_positions = len(position_df)
    total_mtd_pnl = position_df['MTD_PnL'].sum() if not position_df.empty else 0
    latest_day_pnl = position_df['Latest_Day_PnL'].sum() if not position_df.empty else 0
    
    # Top/Bottom performers by route
    if not position_df.empty:
        route_pnl = position_df.groupby(['Route', 'HedgeType'])['MTD_PnL'].sum().reset_index()
        route_pnl = route_pnl.sort_values('MTD_PnL', ascending=False)
        
        top_3 = route_pnl.head(3)
        bottom_3 = route_pnl.tail(3).sort_values('MTD_PnL')
    else:
        top_3 = pd.DataFrame()
        bottom_3 = pd.DataFrame()
    
    summary = f"""FTR Daily Report - {date_formatted}
========================================

📊 POSITION PERFORMANCE
   Total Positions: {total_positions:,}
   Latest Day P&L: ${latest_day_pnl:,.2f}
   MTD P&L: ${total_mtd_pnl:,.2f}

"""
    
    if not top_3.empty:
        summary += "   🟢 Top 3 Performers:\n"
        for _, row in top_3.iterrows():
            summary += f"      {row['Route']} ({row['HedgeType']}): ${row['MTD_PnL']:,.0f}\n"
    
    if not bottom_3.empty:
        summary += "   🔴 Bottom 3 Performers:\n"
        for _, row in bottom_3.iterrows():
            summary += f"      {row['Route']} ({row['HedgeType']}): ${row['MTD_PnL']:,.0f}\n"
    
    summary += f"\n📎 Full report attached: {report_filename}\n"
    
    # Save summary
    summary_path = REPORTS_DIR / "email_summary.txt"
    summary_path.write_text(summary, encoding='utf-8')
    logger.info(f"Email summary saved: {summary_path}")
    
    return summary

    logger.info(f"Positions in snapshot: {len(snapshot_df)}")
    
    # Process data
    position_df = calculate_position_summary(snapshot_df, report_date)
    owner_df = calculate_owner_summary(position_df)
def generate_daily_report():
    """Main function to generate the daily report."""
    logger.info("=" * 50)
    logger.info("FTR Daily Report Generation Started")
    
    # Load data
    snapshot_df, report_date = get_latest_snapshot()
    ledger_df = load_ledger()
    
    logger.info(f"Snapshot date: {report_date}")
    logger.info(f"Positions in snapshot: {len(snapshot_df)}")
    
    # Process data
    position_df = calculate_position_summary(snapshot_df, report_date)
    owner_df = calculate_owner_summary(position_df)
    activity_df = get_recent_activity(ledger_df)
    
    # Generate report
    report_path = create_excel_report(position_df, activity_df, owner_df, report_date)
    
    # Generate email summary
    summary = generate_email_summary(
        position_df, owner_df, report_date, report_path.name
    )
    
    print(summary)
    
    logger.info("Report generation completed")
    return report_path


if __name__ == "__main__":
    generate_daily_report()
